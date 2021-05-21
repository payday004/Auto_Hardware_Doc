import openpyxl
import mailmerge
import os


def main():

    # set up "form.docx" for reading
    template = str(os.getcwd()) + '\\' + "form.docx"

    # set up "input.xlsx" for reading
    input_file = str(os.getcwd()) + '\\' + "input.xlsx"
    wb = openpyxl.load_workbook(filename=input_file)
    ws = wb.active
    
    # setup output folder
    current_dir = os.getcwd()
    output_path = current_dir + '\\' + "output"
    os.mkdir(output_path)

    # get data and write documents
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=2, max_col=ws.max_column):

        # read data
        name = str(row[0].value)
        date = str(row[1].value)
        asset = str(row[2].value)
        serial_num = str(row[3].value)
        hardware_list = str(row[4].value)

        # name format of created documents
        save_string = "hardware_" + name + ".docx"

        # create and save documents
        document = mailmerge.MailMerge(template)
        document.merge(NAME=name, DATE=date[:-8], ASSET=asset, SERIAL_NUM=serial_num, LIST=hardware_list)
        document.write("output" + '\\' + save_string)


if __name__ == '__main__':
    main()
